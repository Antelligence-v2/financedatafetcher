"""
Frontend API wrapper for Streamlit integration.
Provides simplified interface for scraping operations.
"""

import os
import sys
from pathlib import Path
from typing import Optional, Dict, Any, Tuple
from datetime import datetime

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    # Try multiple locations for .env file
    current_dir = Path(__file__).parent.parent.parent
    env_paths = [
        current_dir / ".env",
        current_dir.parent / ".env",
        Path.cwd() / ".env",
    ]
    for env_path in env_paths:
        if env_path.exists():
            load_dotenv(env_path, override=True)  # override=True ensures env vars take precedence
            break
    else:
        # Try default locations
        load_dotenv(override=True)
except ImportError:
    # dotenv not installed, skip
    pass

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.scraper.universal_scraper import UniversalScraper
from src.scraper.dental_etf_scraper import DentalETFScraper
from src.utils.config_manager import ConfigManager
from src.pipeline.pipeline_runner import PipelineRunner, PipelineResult
from src.exporter.excel_exporter import ExcelExporter
from src.utils.logger import get_logger
from src.utils.io_utils import get_output_path, generate_run_id

logger = get_logger()


class FrontendAPI:
    """
    API wrapper for frontend integration.
    Provides simplified methods for scraping operations.
    """
    
    def __init__(self):
        self.config_manager = ConfigManager()
        self.exporter = ExcelExporter()
        self.logger = get_logger()
    
    def scrape_url(
        self,
        url: str,
        use_stealth: bool = True,
        override_robots: bool = False,
        use_fallbacks: bool = True,
    ) -> Dict[str, Any]:
        """
        Scrape data from a URL.
        
        Args:
            url: URL to scrape
            use_stealth: Enable stealth mode
            override_robots: Override robots.txt
            use_fallbacks: Use fallback sources
        
        Returns:
            Dictionary with result data, status, and metadata
        """
        try:
            # Create universal scraper
            scraper = UniversalScraper(
                use_stealth=use_stealth,
                headless=True,
            )
            
            # Run scrape
            result = scraper.scrape_with_discovery(
                url=url,
                override_robots=override_robots,
                save_raw=False,  # Don't save raw files in frontend
            )
            
            if result.success and result.data is not None and not result.data.empty:
                return {
                    "success": True,
                    "data": result.data,
                    "rows": len(result.data),
                    "columns": list(result.data.columns),
                    "warnings": result.validation_warnings,
                    "error": None,
                    "metadata": {
                        "url": url,
                        "source": result.source,
                        "date_range": result.date_range,
                        "run_id": result.run_id,
                    },
                }
            else:
                return {
                    "success": False,
                    "data": None,
                    "rows": 0,
                    "columns": [],
                    "warnings": result.validation_warnings if result.validation_warnings else [],
                    "error": result.error or "No data extracted",
                    "metadata": {
                        "url": url,
                        "source": result.source,
                        "run_id": result.run_id,
                    },
                }
        
        except Exception as e:
            logger.error(f"Error scraping URL: {e}")
            return {
                "success": False,
                "data": None,
                "rows": 0,
                "columns": [],
                "warnings": [],
                "error": str(e),
                "metadata": {
                    "url": url,
                    "source": "unknown",
                    "run_id": generate_run_id(),
                },
            }
    
    def export_to_excel(
        self,
        df,
        filename: Optional[str] = None,
    ) -> Tuple[Optional[bytes], Optional[str]]:
        """
        Export DataFrame to Excel format (in-memory for cloud compatibility).
        
        Args:
            df: DataFrame to export
            filename: Optional filename (without extension)
        
        Returns:
            Tuple of (excel_bytes, filename)
        """
        try:
            if filename is None:
                filename = f"scraped_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Use in-memory export for cloud compatibility
            excel_bytes, filename = self.exporter.export_to_bytes(
                df, 
                filename=filename, 
                site_id=None
            )
            
            return excel_bytes, filename
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None, None
    
    def get_configured_sites(self) -> list:
        """
        Get list of configured sites.
        
        Returns:
            List of site dictionaries with id, name, and page_url
        """
        return self.config_manager.list_sites()
    
    def scrape_configured_site(
        self,
        site_id: str,
        use_stealth: bool = True,
        override_robots: bool = False,
    ) -> Dict[str, Any]:
        """
        Scrape from a configured site.
        
        Args:
            site_id: Site ID from configuration
            use_stealth: Enable stealth mode
            override_robots: Override robots.txt
        
        Returns:
            Dictionary with result data, status, and metadata
        """
        try:
            runner = PipelineRunner(
                config_manager=self.config_manager,
                exporter=self.exporter,
            )
            
            pipeline_result = runner.run(
                site_id=site_id,
                override_robots=override_robots,
                export=False,  # Don't auto-export in frontend
            )
            
            if pipeline_result.success and pipeline_result.data is not None and not pipeline_result.data.empty:
                warnings = []
                if pipeline_result.validation_result:
                    warnings = pipeline_result.validation_result.warnings
                
                return {
                    "success": True,
                    "data": pipeline_result.data,
                    "rows": len(pipeline_result.data),
                    "columns": list(pipeline_result.data.columns),
                    "warnings": warnings,
                    "error": None,
                    "metadata": {
                        "site_id": site_id,
                        "url": pipeline_result.scraper_result.url if pipeline_result.scraper_result else "",
                        "source": pipeline_result.source_used or site_id,
                    },
                }
            else:
                return {
                    "success": False,
                    "data": None,
                    "rows": 0,
                    "columns": [],
                    "warnings": [],
                    "error": pipeline_result.error or "Scraping failed",
                    "metadata": {
                        "site_id": site_id,
                        "source": site_id,
                    },
                }
        
        except Exception as e:
            logger.error(f"Error scraping configured site: {e}")
            return {
                "success": False,
                "data": None,
                "rows": 0,
                "columns": [],
                "warnings": [],
                "error": str(e),
                "metadata": {
                    "site_id": site_id,
                    "source": site_id,
                },
            }
    
    def scrape_dental_source(
        self,
        site_id: str,
        etf_symbol: Optional[str] = None,
        ticker: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Scrape dental ETF data sources with dynamic URL construction.
        
        Args:
            site_id: Site ID from configuration
            etf_symbol: ETF symbol for Yahoo Finance holdings (e.g., "IHI")
            ticker: Stock ticker (deprecated, not currently used)
        
        Returns:
            Dictionary with result data, status, and metadata
        """
        try:
            dental_scraper = DentalETFScraper(
                config_manager=self.config_manager,
                use_stealth=True,
                headless=True,
            )
            
            # Handle Yahoo Finance ETF holdings
            if site_id == "dental_yahoo_etf_holdings" and etf_symbol:
                result = dental_scraper.scrape_yahoo_etf_holdings(etf_symbol)
                return {
                    "success": result.success,
                    "data": result.data,
                    "rows": result.rows,
                    "columns": list(result.data.columns) if result.data is not None else [],
                    "warnings": [],
                    "error": result.error,
                    "metadata": {
                        "site_id": site_id,
                        "symbol": result.symbol,
                        "source": f"Yahoo Finance ({etf_symbol})",
                    },
                }
            
            # Fallback to standard configured site scraping
            else:
                return self.scrape_configured_site(
                    site_id=site_id,
                    use_stealth=True,
                    override_robots=False,
                )
        
        except Exception as e:
            logger.error(f"Error scraping dental source: {e}")
            return {
                "success": False,
                "data": None,
                "rows": 0,
                "columns": [],
                "warnings": [],
                "error": str(e),
                "metadata": {
                    "site_id": site_id,
                    "symbol": etf_symbol or ticker,
                    "source": site_id,
                },
            }
    
    def export_dental_to_excel(
        self,
        dataframes: Dict[str, Any],
        filename: Optional[str] = None,
    ) -> Tuple[Optional[bytes], Optional[str]]:
        """
        Export multiple dental DataFrames to a multi-sheet Excel file.
        
        Args:
            dataframes: Dictionary mapping sheet names to DataFrames
            filename: Optional filename (without extension)
        
        Returns:
            Tuple of (excel_bytes, filename)
        """
        import tempfile
        from pathlib import Path
        import pandas as pd
        
        try:
            if filename is None:
                filename = f"dental_etf_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Ensure .xlsx extension
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
                tmp_path = Path(tmp_file.name)
            
            try:
                # Create Excel writer with multiple sheets
                with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                    for sheet_name, df in dataframes.items():
                        if df is not None and not df.empty:
                            # Truncate sheet name to 31 chars (Excel limit)
                            safe_name = sheet_name[:31]
                            df.to_excel(writer, sheet_name=safe_name, index=False)
                            
                            # Format the worksheet
                            worksheet = writer.sheets[safe_name]
                            self._format_worksheet(worksheet, df)
                
                # Read file as bytes
                with open(tmp_path, "rb") as f:
                    excel_bytes = f.read()
                
                # Clean up temp file
                tmp_path.unlink()
                
                logger.info(f"Exported {len(dataframes)} sheets to memory ({len(excel_bytes)} bytes)")
                return excel_bytes, filename
            
            except Exception as e:
                # Clean up on error
                if tmp_path.exists():
                    tmp_path.unlink()
                raise e
        
        except Exception as e:
            logger.error(f"Error exporting dental data to Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None, None
    
    def _format_worksheet(self, worksheet, df):
        """Apply formatting to a worksheet."""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
        
        # Auto-fit column widths
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 0
            )
            # Cap width at 50 characters
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # Format header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Freeze header row
        worksheet.freeze_panes = "A2"

